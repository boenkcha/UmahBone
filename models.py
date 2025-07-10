import sqlite3
import io
import csv
import pandas as pd
import difflib
from flask import flash
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import re
from collections import defaultdict

class RestoDB:
    def __init__(self, db_path="dbBone.db"):
        self.conn = sqlite3.connect(db_path, check_same_thread=False)
        self.cursor = self.conn.cursor()
        self.create_tables()  # Tambahkan ini

    def create_tables(self):
        self.cursor.executescript("""
            CREATE TABLE IF NOT EXISTS Menu (
                nm_menu TEXT PRIMARY KEY,
                harga REAL,
                porsi INTEGER
            );


            CREATE TABLE IF NOT EXISTS Bahan (
                Satuan TEXT,
                harga REAL
            );

            CREATE TABLE IF NOT EXISTS Resep (
                Nm_menu TEXT,
                Nm_bahan TEXT,
                Qty REAL,
                Satuan_pakai TEXT,
                Konversi REAL,
                Satuan_stok TEXT,
                FOREIGN KEY (Nm_menu) REFERENCES Menu(Nm_menu),
                FOREIGN KEY (Nm_bahan) REFERENCES Stok(Nm_bahan)
            );  

            CREATE TABLE IF NOT EXISTS Paket (
                nm_paket TEXT PRIMARY KEY,
                deskripsi TEXT,
                harga REAL
            );

            CREATE TABLE IF NOT EXISTS PaketDetail (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nm_paket TEXT NOT NULL,
                nama_menu TEXT NOT NULL,
                jumlah INTEGER NOT NULL,
                FOREIGN KEY (nm_paket) REFERENCES Paket(nm_paket)
            );                              

                                  -- Transaksi (header)
            CREATE TABLE IF NOT EXISTS Transaksi (
                no_nota TEXT PRIMARY KEY,
                konsumen TEXT,
                tanggal TEXT,
                jenis_pembayaran TEXT
            );

            -- TransaksiDetail (gabungan menu & paket)
            CREATE TABLE IF NOT EXISTS TransaksiDetail (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                no_nota TEXT,
                jenis TEXT CHECK(jenis IN ('menu', 'paket')),
                nama TEXT,
                jumlah INTEGER,
                harga_satuan INTEGER,
                FOREIGN KEY (no_nota) REFERENCES Transaksi(no_nota)
            );


            -- KebutuhanPaket (hasil breakdown bahan dari paket)
            CREATE TABLE IF NOT EXISTS KebutuhanPaket (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                no_nota TEXT,
                nm_paket TEXT DEFAULT '',
                nm_menu TEXT DEFAULT '',
                nm_bahan TEXT,
                jumlah REAL,
                satuan_pakai TEXT,
                FOREIGN KEY (no_nota) REFERENCES Transaksi(no_nota),
                FOREIGN KEY (nm_bahan) REFERENCES Stok(Nm_bahan)
            );

            CREATE TABLE IF NOT EXISTS Pembelian (
                no_nota TEXT PRIMARY KEY,
                tanggal TEXT,
                nama_supplier TEXT,
                status_pembayaran TEXT
            );

            CREATE TABLE IF NOT EXISTS PembelianDetail (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                no_nota TEXT,
                nm_bahan TEXT,
                jumlah REAL,
                satuan TEXT,
                harga REAL,
                FOREIGN KEY (no_nota) REFERENCES Pembelian(no_nota)
            );
            CREATE TABLE IF NOT EXISTS Stok (
                nm_bahan TEXT PRIMARY KEY,
                jumlah REAL,
                satuan TEXT,
                harga_modal REAL
            );

            CREATE TABLE IF NOT EXISTS NeracaRugiLaba (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bulan TEXT NOT NULL,          -- Format: 'YYYY-MM'
                pendapatan REAL DEFAULT 0,
                hpp REAL DEFAULT 0,           -- Harga Pokok Penjualan
                laba_kotor REAL DEFAULT 0,
                biaya_operasional REAL DEFAULT 0,
                laba_bersih REAL DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
     
            CREATE TABLE IF NOT EXISTS BiayaOperasional (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tanggal_pengeluaran TEXT NOT NULL,  -- Format: 'YYYY-MM-DD'
                keterangan TEXT,
                jumlah REAL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS PelunasanPiutang (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                no_nota TEXT NOT NULL,
                tanggal TEXT NOT NULL,
                jumlah REAL NOT NULL,
                keterangan TEXT,
                metode_pembayaran TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS PelunasanHutang (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                no_nota TEXT,
                tanggal TEXT,
                jumlah REAL,
                keterangan TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS MutasiStok (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tanggal TEXT,
                nm_bahan TEXT,
                satuan TEXT,
                jenis TEXT CHECK(jenis IN ('MASUK', 'KELUAR')),
                jumlah REAL,
                harga_per_unit REAL,
                total_nilai REAL,
                sumber TEXT,
                keterangan TEXT
            );

            CREATE TABLE IF NOT EXISTS konversi_satuan (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nm_bahan TEXT NOT NULL,
                dari_satuan TEXT NOT NULL,
                ke_satuan TEXT NOT NULL,
                faktor_konversi REAL NOT NULL
            );

        """)
        self.conn.commit()

    def get_porsi(self, nm_menu):
        self.cursor.execute("SELECT porsi FROM Menu WHERE nm_menu = ?", (nm_menu,))
        hasil = self.cursor.fetchone()
        return hasil[0] if hasil else ''

    def set_porsi(self, nama_menu, porsi):
        self.cursor.execute("UPDATE Menu SET porsi = ? WHERE nm_menu = ?", (porsi, nama_menu))
        self.conn.commit()

    def hitung_kebutuhan_bahan_dari_paket(self, nm_paket):
        self.cursor.execute("""
            SELECT pd.nama_menu, pd.jumlah, m.porsi
            FROM PaketDetail pd
            JOIN Menu m ON pd.nama_menu = m.nama_menu
            WHERE pd.nm_paket = ?
        """, (nm_paket,))
        isi_paket = self.cursor.fetchall()

        kebutuhan_bahan = {}

        for nama_menu, jumlah_menu, porsi in isi_paket:
            self.cursor.execute("""
                SELECT Nm_bahan, Qty, Konversi
                FROM Resep
                WHERE Nm_menu = ?
            """, (nama_menu,))
            resep_menu = self.cursor.fetchall()

            for bahan, qty, konversi in resep_menu:
                total_qty_stok = (jumlah_menu / porsi) * qty * konversi
                if bahan in kebutuhan_bahan:
                    kebutuhan_bahan[bahan] += total_qty_stok
                else:
                    kebutuhan_bahan[bahan] = total_qty_stok

        return kebutuhan_bahan


    def tambah_menu(self, nama, harga):
        self.cursor.execute("INSERT INTO Menu (Nm_menu, Harga) VALUES (?, ?)", (nama, harga))
        self.conn.commit()

    

    def tambah_resep(self, nm_menu, nm_bahan, qty, satuan_pakai, konversi, satuan_stok):
        self.cursor.execute("""
            INSERT INTO Resep (Nm_menu, Nm_bahan, Qty, Satuan_pakai, Konversi, Satuan_stok)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (nm_menu, nm_bahan, qty, satuan_pakai, konversi, satuan_stok))
        self.conn.commit()


    def tambah_resep_lengkap(self, nm_menu, nm_bahan, qty, satuan_pakai, konversi, satuan_stok):
        self.cursor.execute(
            "INSERT INTO Resep (Nm_menu, Nm_bahan, Qty, Satuan_pakai, Konversi, Satuan_stok) VALUES (?, ?, ?, ?, ?, ?)",
            (nm_menu, nm_bahan, qty, satuan_pakai, konversi, satuan_stok)
        )
        self.conn.commit()

    def tambah_paket(self, nm_paket, nm_menu, harga):
        self.cursor.execute("INSERT INTO Paket (Nm_paket, Nm_menu, Harga) VALUES (?, ?, ?)", (nm_paket, nm_menu, harga))
        self.conn.commit()


    def jual_menu(self, nm_menu, jumlah):
        if not self.cek_stok_cukup(nm_menu, jumlah):
            return False
        self.kurangi_stok(nm_menu, jumlah)
        return True

    def get_menu(self):
        self.cursor.execute("SELECT Nm_menu FROM Menu")
        return [row[0] for row in self.cursor.fetchall()]

    
    def close(self):
        self.conn.close()

    def cek_stok(self, nm_bahan):
        result = self.cursor.execute("""
            SELECT jumlah FROM Stok
            WHERE LOWER(nm_bahan) = LOWER(?)
            ORDER BY jumlah DESC LIMIT 1
        """, (nm_bahan,)).fetchone()
        
        return result[0] if result else None

    def ambil_stok_bahan_dan_satuan(self):
        self.cursor.execute("SELECT nm_bahan, satuan FROM Stok")
        return self.cursor.fetchall()

    def ambil_menu(self):
        self.cursor.execute("SELECT * FROM Menu order by nm_menu")
        return self.cursor.fetchall()

    def ambil_bahan(self):
        self.cursor.execute("""
            SELECT * FROM Bahan 
            ORDER BY nm_bahan COLLATE NOCASE ASC 
            LIMIT 15
        """)
        return self.cursor.fetchall()

    def ambil_semua_stok(self):
        self.cursor.execute("SELECT nm_bahan, satuan FROM Stok ORDER BY nm_bahan")
        return self.cursor.fetchall()

    def ambil_semua_menu(self):
        self.cursor.execute("SELECT Nm_menu FROM Menu")
        return self.cursor.fetchall()

    def ambil_semua_bahan(self):
        self.cursor.execute("SELECT Nm_bahan, Satuan FROM Bahan order by nm_bahan COLLATE NOCASE ASC")
        return self.cursor.fetchall()


    def tambah_resep(self, nm_menu, nm_bahan, qty, satuan_pakai, konversi, satuan_stok):
        self.cursor.execute('''
            INSERT INTO Resep (Nm_menu, Nm_bahan, Qty, Satuan_pakai, Konversi, Satuan_stok)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (nm_menu, nm_bahan, qty, satuan_pakai, konversi, satuan_stok))
        self.conn.commit()

    def get_konversi_data(self):
        self.cursor.execute('SELECT Nm_bahan, Satuan_pakai, Konversi FROM Resep')
        return [
            {'nm_bahan': row[0], 'satuan_pakai': row[1], 'konversi': row[2]}
            for row in self.cursor.fetchall()
        ]

    def ambil_semua_menu_dari_resep(self):
        self.cursor.execute("SELECT DISTINCT nm_menu FROM Resep")
        hasil = [row[0] for row in self.cursor.fetchall()]
        return hasil

    def ambil_resep_per_menu(self, nm_menu):
        self.cursor.execute("""
            SELECT Nm_bahan, Qty, Satuan_pakai, Konversi, Satuan_stok
            FROM Resep
            WHERE Nm_menu = ?
        """, (nm_menu,))
        return self.cursor.fetchall()
    
    def hapus_bahan_dari_resep(self, nm_menu, nm_bahan):
        self.cursor.execute("DELETE FROM Resep WHERE nm_menu = ? AND nm_bahan = ?", (nm_menu, nm_bahan))
        self.conn.commit()


    def tambah_menu_ke_paket(self, nm_paket, nm_menu):
        self.cursor.execute("INSERT INTO Paket (nm_paket, nm_menu) VALUES (?, ?)", (nm_paket, nm_menu))
        self.conn.commit()

    def ambil_semua_paket_dengan_menu(self):
        self.cursor.execute("""
            SELECT nm_paket, GROUP_CONCAT(nm_menu, ', ') AS daftar_menu
            FROM Paket
            GROUP BY nm_paket
        """)
        return self.cursor.fetchall()

    def ambil_semua_menu(self):
        self.cursor.execute("SELECT nm_menu, harga FROM Menu")
        return self.cursor.fetchall()
    
    def ambil_semua_paket(self):
        self.cursor.execute("SELECT nm_paket, deskripsi, harga FROM Paket")
        return self.cursor.fetchall()

    def ambil_detail_paket(self):
        self.cursor.execute("""
            SELECT nm_paket, nm_menu, jumlah FROM PaketDetail
        """)
        return self.cursor.fetchall()
    
    # [(nm_paket, nama_menu, jumlah)]
    def ambil_detail_semua_paket(self):
        self.cursor.execute("""
            SELECT nm_paket, nama_menu, jumlah FROM PaketDetail
        """)
        return self.cursor.fetchall()

    def get_jumlah_porsi_paket(self, nm_paket):
        self.cursor.execute("SELECT jumlah_porsi FROM Paket WHERE nm_paket = ?", (nm_paket,))
        row = self.cursor.fetchone()
        return row[0] if row else None

    def get_porsi_menu(self, nm_menu):
        self.cursor.execute("SELECT porsi FROM Menu WHERE nama_menu = ?", (nm_menu,))
        row = self.cursor.fetchone()
        return row[0] if row else None

    def tambah_menu_ke_paket(self, nm_paket, nama_menu, jumlah):
        self.cursor.execute("INSERT INTO PaketDetail (nm_paket, nama_menu, jumlah) VALUES (?, ?, ?)",
                            (nm_paket, nama_menu, jumlah))
        self.conn.commit()

    def get_all_menu(self):
        cur = self.conn.cursor()
        cur.execute("SELECT nm_menu FROM Menu")
        return [row[0] for row in cur.fetchall()]

    def get_all_paket(self):
        cur = self.conn.cursor()
        cur.execute("SELECT nm_paket FROM Paket")
        return [row[0] for row in cur.fetchall()]

 #MULAI DARI SINI
    def insert_transaksi(self, no_nota, konsumen, tanggal, jenis_pembayaran):
        sql = """
        INSERT INTO Transaksi(no_nota, konsumen, tanggal, jenis_pembayaran)
        VALUES (?, ?, ?, ?)
        """
        self.cursor.execute(sql, (no_nota, konsumen, tanggal, jenis_pembayaran))
       # self.conn.commit()

    def insert_transaksi_detail(self, no_nota, jenis, nama, jumlah, harga_satuan):
        self.conn.execute('''
            INSERT INTO TransaksiDetail (no_nota, jenis, nama, jumlah, harga_satuan)
            VALUES (?, ?, ?, ?, ?)
        ''', (no_nota, jenis, nama, jumlah, harga_satuan))
      #  self.conn.commit()

    #cek ulang
    def hitung_simpan_kebutuhan_bahan(self, no_nota):
        try:
            # Hapus kebutuhan paket lama dulu
            self.cursor.execute("DELETE FROM KebutuhanPaket WHERE no_nota = ?", (no_nota,))

            # Ambil semua item dalam transaksi no_nota
            self.cursor.execute("SELECT jenis, nama, jumlah FROM TransaksiDetail WHERE no_nota = ?", (no_nota,))
            items = self.cursor.fetchall()

            detail_kebutuhan = []  # Untuk insert ke KebutuhanPaket

            # Loop setiap item
            for jenis, nama, jumlah in items:
                if jenis == 'menu':
                    self.cursor.execute("""
                        SELECT Nm_bahan, Qty, Konversi, Satuan_pakai
                        FROM Resep
                        WHERE Nm_menu = ?
                    """, (nama,))
                    resep_list = self.cursor.fetchall()

                    for nm_bahan, qty, konversi, satuan_pakai in resep_list:
                        total_qty = qty * jumlah  # total kebutuhan resep

                        # Potong stok otomatis menggunakan fungsi baru
                        self.potong_stok_otomatis(nm_bahan, total_qty, satuan_pakai)

                        # Ambil harga_modal terbaru setelah potong untuk disimpan
                        self.cursor.execute("""
                            SELECT harga_modal FROM Stok 
                            WHERE nm_bahan = ? LIMIT 1
                        """, (nm_bahan,))
                        harga_modal_row = self.cursor.fetchone()
                        harga_modal = harga_modal_row[0] if harga_modal_row else 0

                        harga_pakai = harga_modal / konversi if konversi else 0
                        detail_kebutuhan.append((no_nota, None, nama, nm_bahan, total_qty, satuan_pakai, harga_pakai))

                elif jenis == 'paket':
                    # Ambil detail paket
                    self.cursor.execute("""
                        SELECT nm_menu, jumlah FROM PaketDetail WHERE nm_paket = ?
                    """, (nama,))
                    paket_detail_list = self.cursor.fetchall()

                    for nama_menu, jumlah_menu in paket_detail_list:
                        self.cursor.execute("""
                            SELECT Nm_bahan, Qty, Konversi, Satuan_pakai
                            FROM Resep
                            WHERE Nm_menu = ?
                        """, (nama_menu,))
                        resep_list = self.cursor.fetchall()

                        for nm_bahan, qty, konversi, satuan_pakai in resep_list:
                            total_qty = qty * jumlah_menu * jumlah

                            # Potong stok otomatis
                            self.potong_stok_otomatis(nm_bahan, total_qty, satuan_pakai)

                            self.cursor.execute("""
                                SELECT harga_modal FROM Stok 
                                WHERE nm_bahan = ? LIMIT 1
                            """, (nm_bahan,))
                            harga_modal_row = self.cursor.fetchone()
                            harga_modal = harga_modal_row[0] if harga_modal_row else 0

                            harga_pakai = harga_modal / konversi if konversi else 0
                            detail_kebutuhan.append((no_nota, nama, nama_menu, nm_bahan, total_qty, satuan_pakai, harga_pakai))

            # Simpan detail kebutuhan bahan ke tabel KebutuhanPaket
            for no_nota, nm_paket, nm_menu, nm_bahan, jumlah, satuan_pakai, harga in detail_kebutuhan:
                self.cursor.execute("""
                    INSERT INTO KebutuhanPaket(no_nota, nm_paket, nm_menu, nm_bahan, jumlah, satuan_pakai, harga)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (no_nota, nm_paket, nm_menu, nm_bahan, jumlah, satuan_pakai, harga))

            self.conn.commit()
            print("✅ Semua proses berhasil. Commit dilakukan.")

        except Exception as e:
            self.conn.rollback()
            print(f"❌ ROLLBACK karena error: {e}")
            raise



    def potong_stok_otomatis(self, nm_bahan, total_qty, satuan_pakai):
        # Cek berapa banyak satuan bahan ini di stok
        self.cursor.execute("""
            SELECT COUNT(DISTINCT satuan) 
            FROM Stok 
            WHERE nm_bahan = ?
        """, (nm_bahan,))
        jumlah_satuan = self.cursor.fetchone()[0]

        if jumlah_satuan == 1:
            # Ambil satuan dan potong langsung dengan konversi di Resep jika ada
            self.cursor.execute("""
                SELECT jumlah, satuan 
                FROM Stok 
                WHERE nm_bahan = ?
            """, (nm_bahan,))
            stok_row = self.cursor.fetchone()
            stok_saat_ini, satuan_stok = stok_row

            # Cek apakah satuan sama, jika tidak gunakan konversi di Resep
            if satuan_pakai == satuan_stok:
                total_qty_stok = total_qty
            else:
                # Ambil faktor konversi dari Resep
                self.cursor.execute("""
                    SELECT Konversi 
                    FROM Resep 
                    WHERE nm_bahan = ? AND Satuan_pakai = ?
                """, (nm_bahan, satuan_pakai))
                konversi_row = self.cursor.fetchone()
                if not konversi_row or not konversi_row[0]:
                    raise Exception(f"Tidak ada konversi di Resep untuk {nm_bahan} dari {satuan_pakai} ke {satuan_stok}")
                faktor_konversi = konversi_row[0]
                total_qty_stok = total_qty / faktor_konversi

            if stok_saat_ini >= total_qty_stok:
                self.cursor.execute("""
                    UPDATE Stok SET jumlah = jumlah - ? 
                    WHERE nm_bahan = ? AND satuan = ?
                """, (total_qty_stok, nm_bahan, satuan_stok))
                print(f"✅ Stok {nm_bahan} ({satuan_stok}) dipotong: {total_qty_stok}")
                return
            else:
                raise Exception(f"Stok {nm_bahan} tidak cukup. Dibutuhkan: {total_qty_stok}, Tersedia: {stok_saat_ini}")

        else:
            # Jika ada lebih dari 1 satuan, gunakan konversi_satuan
            # 1. Cek stok dengan satuan sama dulu
            self.cursor.execute("""
                SELECT harga_modal, jumlah, satuan 
                FROM Stok 
                WHERE nm_bahan = ? AND satuan = ? 
                ORDER BY jumlah DESC
            """, (nm_bahan, satuan_pakai))
            row = self.cursor.fetchone()

            if row and row[1] >= total_qty:
                # Potong langsung
                self.cursor.execute("""
                    UPDATE Stok SET jumlah = jumlah - ? 
                    WHERE nm_bahan = ? AND satuan = ?
                """, (total_qty, nm_bahan, satuan_pakai))
                print(f"✅ Stok {nm_bahan} ({satuan_pakai}) dipotong langsung: {total_qty}")
                return

            # 2. Jika stok tidak cukup atau satuan tidak sama, cari konversi di konversi_satuan
            self.cursor.execute("""
                SELECT s.harga_modal, s.jumlah, s.satuan, k.faktor_konversi 
                FROM Stok s
                JOIN konversi_satuan k 
                ON s.nm_bahan = k.nm_bahan AND s.satuan = k.ke_satuan
                WHERE k.dari_satuan = ? AND s.nm_bahan = ?
                ORDER BY s.jumlah DESC
            """, (satuan_pakai, nm_bahan))
            rows = self.cursor.fetchall()

            for harga_modal, jumlah_stok, satuan_stok, faktor in rows:
                kebutuhan_dalam_satuan_stok = total_qty * faktor
                if jumlah_stok >= kebutuhan_dalam_satuan_stok:
                    self.cursor.execute("""
                        UPDATE Stok SET jumlah = jumlah - ? 
                        WHERE nm_bahan = ? AND satuan = ?
                    """, (kebutuhan_dalam_satuan_stok, nm_bahan, satuan_stok))
                    print(f"✅ Stok {nm_bahan} dikonversi dari {total_qty} {satuan_pakai} menjadi {kebutuhan_dalam_satuan_stok} {satuan_stok} dan dipotong.")
                    return

            # 3. Jika semua stok tidak cukup
            raise Exception(f"❌ Stok {nm_bahan} tidak cukup di semua satuan.")
      

    # Contoh method untuk mengambil daftar menu dan paket untuk dropdown
    def get_daftar_menu(self):
        self.cursor.execute("SELECT nm_menu, harga FROM Menu ORDER BY nm_menu")
        return [{'nama': row[0], 'harga': row[1]} for row in self.cursor.fetchall()]

    def get_daftar_paket(self):
        self.cursor.execute("SELECT nm_paket, harga FROM Paket ORDER BY nm_paket")
        return [{'nama': row[0], 'harga': row[1]} for row in self.cursor.fetchall()]

    def get_resep_menu(self, nm_menu):
        self.cursor.execute("SELECT Nm_bahan, Qty, Konversi, Satuan_pakai FROM Resep WHERE Nm_menu = ?", (nm_menu,))
        return self.cursor.fetchall()

    def get_menu_dalam_paket(self, nm_paket):
        self.cursor.execute("SELECT nm_menu, jumlah FROM PaketDetail WHERE nm_paket = ?", (nm_paket,))
        return self.cursor.fetchall()
    

    def hapus_menu(self, nama_menu):
        conn = self.conn
        cursor = conn.cursor()

        # Hapus dari tabel Resep terlebih dahulu untuk menjaga integritas referensial
        cursor.execute("DELETE FROM Resep WHERE Nm_menu = ?", (nama_menu,))

        # Hapus dari tabel PaketDetail (jika ada paket yang menggunakan menu ini)
        cursor.execute("DELETE FROM PaketDetail WHERE nm_menu = ?", (nama_menu,))

        # Hapus dari tabel Menu
        cursor.execute("DELETE FROM Menu WHERE nm_menu = ?", (nama_menu,))

        conn.commit()


    def get_transaksi(self, no_nota):
        self.cursor.execute("""
            SELECT no_nota, konsumen, tanggal, jenis_pembayaran
            FROM Transaksi WHERE no_nota = ?
        """, (no_nota,))
        return self.cursor.fetchone()

    def get_transaksi_by_nota(self, no_nota):
        self.cursor.execute("""
            SELECT no_nota, konsumen, tanggal, jenis_pembayaran
            FROM Transaksi WHERE no_nota = ?
        """, (no_nota,))
        return self.cursor.fetchone()

    def get_transaksi_detail(self, no_nota):
        self.cursor.execute("""
            SELECT jenis, nama, jumlah, harga_satuan
            FROM TransaksiDetail WHERE no_nota = ?
        """, (no_nota,))
        return self.cursor.fetchall()

    def get_kebutuhan_bahan(self, no_nota):
        self.cursor.execute("""
            SELECT 
                nm_bahan,
                jumlah,
                satuan_pakai,
                harga -- langsung ambil dari tabel KebutuhanPaket
            FROM KebutuhanPaket
            WHERE no_nota = ?
        """, (no_nota,))
        
        result = []
        for nama, jumlah, satuan_pakai, harga_per_satuan_pakai in self.cursor.fetchall():
            total = jumlah * harga_per_satuan_pakai
            result.append((nama, jumlah, satuan_pakai, harga_per_satuan_pakai, total))
        return result

    def get_stok_list(self, limit=10, offset=0):
        self.cursor.execute("""
            SELECT Nm_bahan, jumlah, satuan, harga_modal 
            FROM Stok
            ORDER BY LOWER(Nm_bahan) 
            LIMIT ? OFFSET ?
        """, (limit, offset))
        return self.cursor.fetchall()

    def search_stok_by_nama(self, nama_bahan, limit=10, offset=0):
        self.cursor.execute("""
            SELECT Nm_bahan, jumlah, satuan, harga_modal 
            FROM Stok 
            WHERE Nm_bahan LIKE ? 
            LIMIT ? OFFSET ?
        """, (f"%{nama_bahan}%", limit, offset))
        return self.cursor.fetchall()

    def count_stok_by_nama(self, nama_bahan):
        self.cursor.execute("""
            SELECT COUNT(*) 
            FROM Stok 
            WHERE Nm_bahan LIKE ?
        """, (f"%{nama_bahan}%",))
        return self.cursor.fetchone()[0]

    def count_stok(self):
        self.cursor.execute("SELECT COUNT(*) FROM Stok")
        return self.cursor.fetchone()[0]
    
    def get_satuan_bahan(self, nama_bahan):
        self.cursor.execute("SELECT Satuan FROM Bahan WHERE Nm_bahan = ?", (nama_bahan,))
        row = self.cursor.fetchone()
        return row[0] if row else ''

    def get_daftar_bahan(self):
        self.cursor.execute("SELECT Nm_bahan FROM Bahan ORDER BY Nm_bahan")
        return [row[0] for row in self.cursor.fetchall()]


    def get_transaksi_by_tanggal(self, tanggal_awal, tanggal_akhir):
        self.cursor.execute("""
            SELECT no_nota, konsumen, tanggal, jenis_pembayaran
            FROM Transaksi
            WHERE tanggal BETWEEN ? AND ?
            ORDER BY tanggal
        """, (tanggal_awal, tanggal_akhir))
        return self.cursor.fetchall()
    
    def get_riwayat_pembelian(self, limit=10):
        self.cursor.execute("""
            SELECT p.no_nota, p.tanggal, p.nama_supplier, p.status_pembayaran,
                d.nm_bahan, d.jumlah, d.satuan, d.harga
            FROM Pembelian p
            JOIN PembelianDetail d ON p.no_nota = d.no_nota
            ORDER BY p.tanggal DESC, p.no_nota DESC
            LIMIT ?
        """, (limit,))
        return self.cursor.fetchall()

    def get_pembelian_by_nota(self, no_nota):
        self.cursor.execute("""
            SELECT p.no_nota, p.tanggal, p.nama_supplier, p.status_pembayaran,
                d.nm_bahan, d.jumlah, d.satuan, d.harga
            FROM Pembelian p
            JOIN PembelianDetail d ON p.no_nota = d.no_nota
            WHERE p.no_nota = ?
        """, (no_nota,))
        return self.cursor.fetchall()


    def insert_pembelian(self, no_nota, tanggal, nama_supplier, status_pembayaran):
        # Cek apakah no_nota sudah ada
        self.cursor.execute("SELECT 1 FROM Pembelian WHERE no_nota = ?", (no_nota,))
        if self.cursor.fetchone():
            raise ValueError(f"No nota '{no_nota}' sudah pernah digunakan.")

        # Pastikan tanggal valid ISO (YYYY-MM-DD)
        try:
            datetime.strptime(tanggal, "%Y-%m-%d")  # validasi format, tanpa ubah
        except ValueError:
            raise ValueError("Format tanggal tidak valid. Gunakan format YYYY-MM-DD.")

        # Insert ke database
        self.cursor.execute("""
            INSERT INTO Pembelian (no_nota, tanggal, nama_supplier, status_pembayaran)
            VALUES (?, ?, ?, ?)
        """, (no_nota, tanggal, nama_supplier, status_pembayaran))
        self.conn.commit()

    def insert_pembelian_detail(self, no_nota, nm_bahan, jumlah, satuan, harga):
        self.cursor.execute("""
            INSERT INTO PembelianDetail (no_nota, nm_bahan, jumlah, satuan, harga)
            VALUES (?, ?, ?, ?, ?)
        """, (no_nota, nm_bahan, jumlah, satuan, harga))
        self.conn.commit()

    def update_stok_setelah_pembelian(self, nm_bahan, jumlah_baru, satuan, harga_total_baru, catat_mutasi=True):
        if not nm_bahan:
            raise ValueError("Nama bahan tidak boleh kosong")

        if jumlah_baru == 0:
            return  # Tidak ada perubahan

        harga_per_unit_baru = harga_total_baru / abs(jumlah_baru) if jumlah_baru != 0 else 0

        # Ambil stok lama
        self.cursor.execute("""
            SELECT jumlah, harga_modal FROM Stok
            WHERE nm_bahan = ? AND satuan = ?
        """, (nm_bahan, satuan))
        row = self.cursor.fetchone()

        jumlah_sebelumnya = row[0] if row else 0
        jumlah_setelah = jumlah_sebelumnya + jumlah_baru

        if jumlah_setelah < 0:
            raise ValueError(f"Stok tidak cukup. Stok sekarang: {jumlah_sebelumnya}, pengurangan: {abs(jumlah_baru)}")

        if row:
            if jumlah_setelah == 0:
                self.cursor.execute("""
                    UPDATE Stok SET jumlah = 0
                    WHERE nm_bahan = ? AND satuan = ?
                """, (nm_bahan, satuan))
            else:
                total_nilai_lama = jumlah_sebelumnya * row[1]
                total_nilai_baru = jumlah_baru * harga_per_unit_baru
                harga_modal_baru = (total_nilai_lama + total_nilai_baru) / jumlah_setelah
                self.cursor.execute("""
                    UPDATE Stok SET jumlah = ?, harga_modal = ?
                    WHERE nm_bahan = ? AND satuan = ?
                """, (jumlah_setelah, harga_modal_baru, nm_bahan, satuan))
        else:
            if jumlah_baru > 0:
                harga_modal_baru = harga_total_baru / jumlah_baru
                self.cursor.execute("""
                    INSERT INTO Stok (nm_bahan, jumlah, satuan, harga_modal)
                    VALUES (?, ?, ?, ?)
                """, (nm_bahan, jumlah_baru, satuan, harga_modal_baru))
            else:
                raise ValueError("Tidak bisa mengurangi stok karena belum tersedia.")

        self.conn.commit()

        # ✅ Catat mutasi HANYA jika jumlah berubah secara signifikan
        if catat_mutasi and jumlah_baru != 0:
            jenis = 'MASUK' if jumlah_baru > 0 else 'KELUAR'
            sumber = 'Pembelian' if jumlah_baru > 0 else 'Koreksi'
            keterangan = 'Pembelian bahan' if jumlah_baru > 0 else 'Koreksi stok pembelian'

            self.catat_mutasi_stok(
                nm_bahan=nm_bahan,
                satuan=satuan,
                jenis=jenis,
                jumlah=abs(jumlah_baru),
                harga_per_unit=harga_per_unit_baru,
                sumber=sumber,
                keterangan=keterangan
            )

    def delete_pembelian_detail_bahan(self, no_nota, nm_bahan, satuan):
        self.cursor.execute("""
            DELETE FROM PembelianDetail
            WHERE no_nota = ? AND nm_bahan = ? AND satuan = ?
        """, (no_nota, nm_bahan, satuan))
        self.conn.commit()


    def hapus_pembelian_jika_detail_kosong(self, no_nota):
        self.cursor.execute("SELECT COUNT(*) FROM PembelianDetail WHERE no_nota = ?", (no_nota,))
        jumlah_detail = self.cursor.fetchone()[0]

        if jumlah_detail == 0:
            self.cursor.execute("DELETE FROM Pembelian WHERE no_nota = ?", (no_nota,))
            self.conn.commit()

    def catat_mutasi_stok(self, nm_bahan, satuan, jenis, jumlah, harga_per_unit, sumber, keterangan):
        from datetime import datetime

        total_nilai = jumlah * harga_per_unit
        tanggal = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        self.cursor.execute("""
            INSERT INTO MutasiStok (tanggal, nm_bahan, satuan, jenis, jumlah, harga_per_unit, total_nilai, sumber, keterangan)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (tanggal, nm_bahan, satuan, jenis, jumlah, harga_per_unit, total_nilai, sumber, keterangan))

        self.conn.commit()


    def update_pembelian(self, no_nota, tanggal, nama_supplier, status_pembayaran):
        self.cursor.execute("""
            UPDATE Pembelian
            SET tanggal = ?, nama_supplier = ?, status_pembayaran = ?
            WHERE no_nota = ?
        """, (tanggal, nama_supplier, status_pembayaran, no_nota))
        self.conn.commit()

    def delete_pembelian_detail(self, no_nota):
        self.cursor.execute("DELETE FROM PembelianDetail WHERE no_nota = ?", (no_nota,))
        self.conn.commit()

    def import_pembelian_csv(self, file):
        text = file.stream.read().decode('utf-8')
        stream = io.StringIO(text)
        reader = csv.DictReader(stream)
        inserted_nota = set()

        for row in reader:
            try:
                tanggal = row['Tanggal']
                no_nota = row['No_Nota'].strip()
                supplier = row['nama_suplier'].strip()
                status = row['status_pembayaran'].strip()

                # === 🧽 Normalisasi nama bahan ===
                nm_bahan = ' '.join(row['Nm_Barang'].strip().upper().split())
                jumlah = float(row['Qty'])
                satuan = row['Satuan'].strip()
                harga = float(row['Harga_modal'])

                # === 🔄 Cek alias bahan (disamakan juga dengan TRIM + UPPER)
                cur = self.conn.cursor()
                cur.execute("""
                    SELECT nama_standar FROM BahanAlias
                    WHERE UPPER(TRIM(nama_lama)) = ?
                """, (nm_bahan,))
                alias_row = cur.fetchone()
                if alias_row:
                    nm_bahan = alias_row[0]  # Gunakan nama standar

                # === ✅ Simpan header pembelian jika belum pernah
                if no_nota not in inserted_nota:
                    existing = self.get_pembelian_by_nota(no_nota)
                    if existing:
                        flash(f'No Nota {no_nota} sudah ada di database, data diabaikan', 'warning')
                        continue
                    self.insert_pembelian(no_nota, tanggal, supplier, status)
                    inserted_nota.add(no_nota)

                # === 🧾 Simpan detail pembelian & update stok
                self.insert_pembelian_detail(no_nota, nm_bahan, jumlah, satuan, harga)
                total_harga = jumlah * harga
                self.update_stok_setelah_pembelian(nm_bahan, jumlah, satuan, total_harga)

                # === 📦 Update atau tambah bahan
                existing_bahan = self.get_bahan_by_nama(nm_bahan)
                if existing_bahan:
                    self.update_harga_bahan(nm_bahan, harga)
                else:
                    self.tambah_bahan(nm_bahan, satuan, harga)

            except Exception as e:
                flash(f'Gagal proses baris: No Nota {row.get("No_Nota", "")}, Bahan {row.get("Nm_Barang", "")} — {str(e)}', 'danger')
                continue

        flash('Proses import selesai.', 'success')



    def get_bahan_by_nama(self, nm_bahan):
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM Bahan WHERE Nm_bahan = ?", (nm_bahan,))
        return cur.fetchone()
    
    def update_harga_bahan(self, nm_bahan, harga):
        cur = self.conn.cursor()
        cur.execute("UPDATE Bahan SET Harga = ? WHERE Nm_bahan = ?", (harga, nm_bahan))
        self.conn.commit()

    def tambah_bahan(self, nm_bahan, satuan, harga):
        cur = self.conn.cursor()
        cur.execute("INSERT INTO Bahan (Nm_bahan, Satuan, Harga) VALUES (?, ?, ?)", (nm_bahan, satuan, harga))
        self.conn.commit()
   
    # Mengecek apakah bahan + satuan sudah ada
    def cek_bahan_ada(self, nm_bahan, satuan):
        cur = self.conn.cursor()
        cur.execute("SELECT 1 FROM Bahan WHERE Nm_bahan = ? AND Satuan = ?", (nm_bahan, satuan))
        return cur.fetchone() is not None

    # Menambahkan bahan baru
    def tambah_bahan(self, nm_bahan, satuan, harga):
        cur = self.conn.cursor()
        cur.execute("INSERT INTO Bahan (Nm_bahan, Satuan, Harga) VALUES (?, ?, ?)", (nm_bahan, satuan, harga))
        self.conn.commit()


    def get_kartu_stok(self):
        # Ambil konversi dan satuan stok
        konversi = {}
        satuan_stok = {}
        self.cursor.execute("SELECT DISTINCT nm_bahan, konversi, satuan_stok FROM Resep WHERE konversi > 0")
        for row in self.cursor.fetchall():
            konversi[row[0]] = row[1]
            satuan_stok[row[0]] = row[2]

        # Pembelian (masuk stok) - tidak perlu konversi
        self.cursor.execute("""
            SELECT p.tanggal, d.nm_bahan, d.jumlah, d.satuan
            FROM PembelianDetail d
            JOIN Pembelian p ON d.no_nota = p.no_nota
        """)
        pembelian = pd.DataFrame(self.cursor.fetchall(), columns=['tanggal', 'nm_bahan', 'jumlah', 'satuan'])
        if not pembelian.empty:
            pembelian['tanggal'] = pd.to_datetime(pembelian['tanggal']).dt.strftime('%Y-%m-%d')
            pembelian['jumlah_stok'] = pembelian['jumlah']
            pembelian['tipe'] = 'masuk'
            pembelian['satuan_stok'] = pembelian['nm_bahan'].map(lambda b: satuan_stok.get(b, pembelian.loc[pembelian['nm_bahan'] == b, 'satuan'].values[0]))

        # Kebutuhan (keluar stok) - perlu konversi balik
        self.cursor.execute("""
            SELECT t.tanggal, k.nm_bahan, k.jumlah, k.satuan_pakai
            FROM KebutuhanPaket k
            JOIN Transaksi t ON k.no_nota = t.no_nota
        """)
        kebutuhan = pd.DataFrame(self.cursor.fetchall(), columns=['tanggal', 'nm_bahan', 'jumlah_pakai', 'satuan_pakai'])
        if not kebutuhan.empty:
            kebutuhan['tanggal'] = pd.to_datetime(kebutuhan['tanggal']).dt.strftime('%Y-%m-%d')
            kebutuhan['jumlah_stok'] = kebutuhan.apply(
                lambda row: -(row['jumlah_pakai'] / konversi.get(row['nm_bahan'], 1)) if row['nm_bahan'] in konversi else -row['jumlah_pakai'],
                axis=1
            )
            kebutuhan['tipe'] = 'keluar'
            kebutuhan['satuan_stok'] = kebutuhan['nm_bahan'].map(lambda b: satuan_stok.get(b, kebutuhan.loc[kebutuhan['nm_bahan'] == b, 'satuan_pakai'].values[0]))

        # Gabungkan
        df = pd.concat([pembelian, kebutuhan], ignore_index=True)
        if df.empty:
            return pd.DataFrame()

        df = df[['tanggal', 'nm_bahan', 'jumlah_stok', 'satuan_stok']]
        df = df.groupby(['nm_bahan', 'satuan_stok', 'tanggal'])['jumlah_stok'].sum().reset_index()

        # Susun laporan
        bahan_list = df['nm_bahan'].unique()
        tanggal_list = sorted(df['tanggal'].unique())
        result = {}

        for bahan in bahan_list:
            satuan = df[df['nm_bahan'] == bahan]['satuan_stok'].iloc[0]
            saldo = 0
            row = {}
            df_bahan = df[df['nm_bahan'] == bahan].set_index('tanggal')['jumlah_stok'].to_dict()
            for tgl in tanggal_list:
                perubahan = df_bahan.get(tgl, 0)
                awal = saldo
                saldo += perubahan
                if perubahan > 0:
                    log = f"+{perubahan:.2f}"
                elif perubahan < 0:
                    log = f"-{abs(perubahan):.2f}"
                else:
                    log = ""
                row[tgl] = f"{awal:.2f} {log} = {saldo:.2f} {satuan}" if log else f"{saldo:.2f} {satuan}"
            row['Saldo Akhir'] = f"{saldo:.2f} {satuan}"
            result[bahan] = row

        report_df = pd.DataFrame(result).T.fillna('')
        report_df = report_df[tanggal_list + ['Saldo Akhir']]
        report_df.index.name = 'Bahan'
        return report_df.reset_index()

    #UNTUK WARNA MASUK DAN KELUAR imi

    def render_html_kartu_stok(self, start=None, end=None):
        df = self.get_kartu_stok_formatted(start, end)
        if df.empty:
            return "<p>Tidak ada data.</p>"

        # Ubah kolom menjadi MultiIndex
        new_columns = []
        for col in df.columns:
            if "_" in col and col not in ["Saldo"]:
                tgl, label = col.split("_")
                new_columns.append((tgl, label))
            else:
                new_columns.append(("", col))
        df.columns = pd.MultiIndex.from_tuples(new_columns)

        # Format angka
        fmt_dict = {
            col: "{:.2f}" for col in df.columns
            if col[1] not in ["Bahan", "Saldo"]
        }

        # Base highlight kolom (SA & M)
        def highlight_column(col):
            if isinstance(col.name, tuple):
                if col.name[1] == "SA":
                    return ["background-color: #d0e7ff"] * len(col)
                elif col.name[1] == "M":
                    return ["background-color: #d0ffd6"] * len(col)
            return [""] * len(col)

        # Tambahan gaya per sel (warna teks biru/ungu tebal)
        def color_cells(val, col_label):
            if col_label[1] == "M" and val != 0:
                return "color: blue; font-weight: bold;"
            elif col_label[1] == "K" and val != 0:
                return "color: red; font-weight: bold;"
            return ""

        # Bangun DataFrame CSS style per sel
        styles = pd.DataFrame("", index=df.index, columns=df.columns)
        for col in df.columns:
            styles[col] = df[col].apply(lambda val: color_cells(val, col))

        styled = (
            df.style
            .apply(highlight_column, axis=0)     # Highlight latar
            .apply(lambda _: styles, axis=None)  # Highlight per sel
            .format(fmt_dict)                    # Format angka
            .set_table_styles([
                {'selector': 'th, td', 'props': [('border', '1px solid black'), ('padding', '5px')]},
                {'selector': 'table', 'props': [('border-collapse', 'collapse'), ('width', '100%')]}
            ])
        )

        return styled.to_html(index=False, escape=False)



    def get_kartu_stok_formatted(self, start=None, end=None):
        # Ambil konversi dan satuan stok
        konversi = {}
        satuan_stok = {}
        self.cursor.execute("SELECT DISTINCT nm_bahan, konversi, satuan_stok FROM Resep WHERE konversi > 0")
        for row in self.cursor.fetchall():
            konversi[row[0]] = row[1]
            satuan_stok[row[0]] = row[2]

        # Ambil pembelian
        self.cursor.execute("""
            SELECT p.tanggal, d.nm_bahan, d.jumlah, d.satuan
            FROM PembelianDetail d
            JOIN Pembelian p ON d.no_nota = p.no_nota
        """)
        pembelian = pd.DataFrame(self.cursor.fetchall(), columns=['tanggal', 'nm_bahan', 'jumlah', 'satuan'])

        # Ambil kebutuhan
        self.cursor.execute("""
            SELECT t.tanggal, k.nm_bahan, k.jumlah, k.satuan_pakai
            FROM KebutuhanPaket k
            JOIN Transaksi t ON k.no_nota = t.no_nota
        """)
        kebutuhan = pd.DataFrame(self.cursor.fetchall(), columns=['tanggal', 'nm_bahan', 'jumlah', 'satuan'])

        if pembelian.empty and kebutuhan.empty:
            return pd.DataFrame()

        if not pembelian.empty:
            pembelian['tanggal'] = pd.to_datetime(pembelian['tanggal'])
            pembelian['masuk'] = pembelian['jumlah']
            pembelian['keluar'] = 0.0
            pembelian['satuan_stok'] = pembelian['nm_bahan'].map(lambda b: satuan_stok.get(b, pembelian.loc[pembelian['nm_bahan'] == b, 'satuan'].values[0]))

        if not kebutuhan.empty:
            kebutuhan['tanggal'] = pd.to_datetime(kebutuhan['tanggal'])
            kebutuhan['masuk'] = 0.0
            kebutuhan['keluar'] = kebutuhan.apply(
                lambda row: (row['jumlah'] / konversi.get(row['nm_bahan'], 1)) if row['nm_bahan'] in konversi else row['jumlah'],
                axis=1
            )
            kebutuhan['satuan_stok'] = kebutuhan['nm_bahan'].map(lambda b: satuan_stok.get(b, kebutuhan.loc[kebutuhan['nm_bahan'] == b, 'satuan'].values[0]))

        df = pd.concat([pembelian, kebutuhan], ignore_index=True)

        # Hitung saldo awal (sebelum start)
        df['tanggal'] = pd.to_datetime(df['tanggal'])
        if start:
            start_date = pd.to_datetime(start)
            saldo_awal_df = df[df['tanggal'] < start_date].groupby(['nm_bahan', 'satuan_stok'])[['masuk', 'keluar']].sum().reset_index()
            saldo_awal_map = {
                (row['nm_bahan'], row['satuan_stok']): row['masuk'] - row['keluar']
                for _, row in saldo_awal_df.iterrows()
            }
            df = df[df['tanggal'] >= start_date]
        else:
            saldo_awal_map = {}

        if end:
            end_date = pd.to_datetime(end)
            df = df[df['tanggal'] <= end_date]

        if df.empty:
            return pd.DataFrame()

        df['tanggal'] = df['tanggal'].dt.strftime('%Y-%m-%d')
        df_sum = df.groupby(['nm_bahan', 'satuan_stok', 'tanggal'])[['masuk', 'keluar']].sum().reset_index()
        bahan_list = df_sum['nm_bahan'].unique()
        tanggal_list = sorted(df_sum['tanggal'].unique())

        hasil = []
        for bahan in bahan_list:
            satuan = df_sum[df_sum['nm_bahan'] == bahan]['satuan_stok'].iloc[0]
            saldo = saldo_awal_map.get((bahan, satuan), 0)
            row = {'Bahan': bahan}
            for tgl in tanggal_list:
                data = df_sum[(df_sum['nm_bahan'] == bahan) & (df_sum['tanggal'] == tgl)]
                masuk = data['masuk'].iloc[0] if not data.empty else 0
                keluar = data['keluar'].iloc[0] if not data.empty else 0
                row[f"{tgl}_SA"] = round(saldo, 2)
                row[f"{tgl}_M"] = masuk
                row[f"{tgl}_K"] = keluar
                saldo += masuk - keluar
            row['Saldo'] = f"{saldo:.2f} {satuan}"
            hasil.append(row)

        # Urutkan kolom
        columns = ['Bahan']
        for tgl in tanggal_list:
            columns += [f"{tgl}_SA", f"{tgl}_M", f"{tgl}_K"]
        columns += ['Saldo']

        return pd.DataFrame(hasil)[columns]
    
    def export_kartu_stok_excel_buffer(self, start=None, end=None):
        df = self.get_kartu_stok_formatted(start, end)
        if df.empty:
            return None

        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Kartu Stok"

        # Ambil list tanggal dari kolom df
        tanggal_set = sorted(set(col.split("_")[0] for col in df.columns if "_" in col and col != 'Saldo'))
        header_row_1 = ['Bahan']
        header_row_2 = ['']
        for tgl in tanggal_set:
            header_row_1.extend([tgl, '', ''])
            header_row_2.extend(['SA', 'M', 'K'])
        header_row_1.append("Saldo")
        header_row_2.append("")

        # Tulis header ke worksheet
        ws.append(header_row_1)
        ws.append(header_row_2)

        # Merge cell untuk header baris pertama
        col_index = 2
        for _ in tanggal_set:
            ws.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=col_index + 2)
            col_index += 3
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.merge_cells(start_row=1, start_column=col_index, end_row=2, end_column=col_index)

        # Styling
        bold_font = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center')
        border_style = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.font = bold_font
                cell.alignment = center
                cell.border = border_style

        # Tulis data
        for idx, (_, row) in enumerate(df.iterrows(), start=3):
            values = [row['Bahan']]
            for tgl in tanggal_set:
                values.append(row.get(f"{tgl}_SA", ""))
                values.append(row.get(f"{tgl}_M", ""))
                values.append(row.get(f"{tgl}_K", ""))
            values.append(row['Saldo'])
            ws.append(values)

            for col in range(1, len(values) + 1):
                cell = ws.cell(row=idx, column=col)
                cell.border = border_style
                cell.alignment = left if col == 1 else center

        # Auto-width untuk semua kolom
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        wb.save(output)
        output.seek(0)
        return output


    def get_pemasukan_bahan(self, nm_bahan):
        self.cursor.execute("""
            SELECT 
                p.tanggal,
                pd.jumlah,
                p.nama_supplier,
                pd.satuan,
                pd.harga
            FROM PembelianDetail pd
            JOIN Pembelian p ON p.no_nota = pd.no_nota
            WHERE pd.nm_bahan = ?
        """, (nm_bahan,))
        return self.cursor.fetchall()

    def get_pengeluaran_bahan(self, nm_bahan):
        self.cursor.execute("""
            SELECT 
                t.tanggal,
                k.jumlah,
                k.satuan_pakai
            FROM KebutuhanPaket k
            JOIN Transaksi t ON t.no_nota = k.no_nota
            WHERE k.nm_bahan = ?
        """, (nm_bahan,))
        return self.cursor.fetchall()


    def get_stokcard(self, nm_bahan):
        # Ambil data pemasukan dan pengeluaran
        pemasukan = self.get_pemasukan_bahan(nm_bahan)
        pengeluaran = self.get_pengeluaran_bahan(nm_bahan)

        # Bentuk list baru dengan format seragam, tambah keterangan dan pisahkan kolom masuk/keluar
        daftar = []

        for p in pemasukan:
            daftar.append({
                "tanggal": p[0],           # tanggal
                "keterangan": f"Pembelian {p[2]}",  # nama_supplier
                "masuk": p[1],             # jumlah
                "keluar": 0,
                "satuan": p[3],            # satuan
                "harga_modal": p[4],       # harga
            })

        for q in pengeluaran:
            daftar.append({
                "tanggal": q[0],           # tanggal
                "keterangan": "Pengeluaran (Transaksi)",  # bisa dikembangkan jika ada info lain
                "masuk": 0,
                "keluar": q[1],            # jumlah
                "satuan": q[2],            # satuan_pakai
                "harga_modal": None,
            })

        # Urutkan berdasarkan tanggal (asumsi format 'YYYY-MM-DD' cocok untuk sort string)
        daftar.sort(key=lambda x: x['tanggal'])

        return daftar

    def get_info_stok(self, nm_bahan):
        cursor = self.conn.cursor()

        # 1. Ambil semua alias yang mengarah ke nm_bahan
        alias_rows = cursor.execute("""
            SELECT nama_lama FROM BahanAlias 
            WHERE LOWER(nama_standar) = LOWER(?)
        """, (nm_bahan,)).fetchall()
        alias_list = [row[0] for row in alias_rows]
        semua_nama = [nm_bahan] + alias_list

        # 2. Ambil pemasukan dari semua nama
        format_q = ",".join("?" for _ in semua_nama)
        pemasukan = cursor.execute(f"""
            SELECT tanggal, no_nota, jumlah, satuan, harga, nm_bahan
            FROM Pembelian JOIN PembelianDetail USING(no_nota)
            WHERE LOWER(nm_bahan) IN ({format_q})
            ORDER BY tanggal
        """, [n.lower() for n in semua_nama]).fetchall()

        # 3. Ambil pengeluaran dari semua nama
        pengeluaran = cursor.execute(f"""
            SELECT t.tanggal, kp.no_nota, kp.jumlah, kp.satuan_pakai, kp.nm_bahan
            FROM KebutuhanPaket kp
            JOIN Transaksi t ON t.no_nota = kp.no_nota
            WHERE LOWER(kp.nm_bahan) IN ({format_q})
            ORDER BY t.tanggal
        """, [n.lower() for n in semua_nama]).fetchall()

        # 4. Ambil satuan stok utama dari tabel Stok
        stok_info = cursor.execute("""
            SELECT satuan FROM Stok 
            WHERE LOWER(nm_bahan) = LOWER(?)
        """, (nm_bahan,)).fetchone()
        satuan_stok = stok_info[0] if stok_info else ""

        # 5. Ambil konversi satuan (jika ada)
        konversi_dict = {}
        hasil_konv = cursor.execute("""
            SELECT satuan_pakai, satuan_stok, konversi
            FROM Resep 
            WHERE LOWER(nm_bahan) = LOWER(?)
        """, (nm_bahan,)).fetchall()
        for satuan_pakai, satuan_stok_db, konversi in hasil_konv:
            if satuan_stok_db.lower() == satuan_stok.lower():
                konversi_dict[satuan_pakai.lower()] = konversi

        # 6. Gabungkan dan kelompokkan berdasarkan satuan
        grouped = defaultdict(list)
        saldo_map = defaultdict(float)

        for tgl, nota, jml, satuan, harga, asal in pemasukan:
            satuan_key = satuan.lower()
            nilai = jml * harga
            saldo_map[satuan_key] += jml
            grouped[satuan_key].append({
                "tanggal": tgl,
                "no_nota": nota,
                "jenis": "Masuk",
                "jumlah": jml,
                "satuan": satuan,
                "nilai": nilai,
                "saldo": saldo_map[satuan_key],
                "satuan_stok": satuan,
                "asal": asal
            })

        for tgl, nota, jml, satuan_pakai, asal in pengeluaran:
            satuan_key = satuan_pakai.lower()
            konversi = konversi_dict.get(satuan_key)
            jumlah_stok = jml / konversi if konversi else jml
            saldo_map[satuan_key] -= jumlah_stok
            grouped[satuan_key].append({
                "tanggal": tgl,
                "no_nota": nota,
                "jenis": "Keluar",
                "jumlah": jml,
                "satuan": satuan_pakai,
                "nilai": "-",
                "saldo": saldo_map[satuan_key],
                "satuan_stok": satuan_pakai,
                "asal": asal
            })

        # Urutkan tiap grup berdasarkan tanggal
        for satuan_key in grouped:
            grouped[satuan_key].sort(key=lambda x: x["tanggal"])

        return {
            "records_by_satuan": grouped,      # key: satuan.lower(), value: list transaksi
            "alias_list": alias_list,          # daftar alias nama_lama
            "satuan_stok": satuan_stok         # satuan utama stok
        }




    def eksekusi_select(self, query, params=()):
            cur = self.conn.cursor()
            cur.execute(query, params)
            return cur.fetchall()
    
    def normalisasi_nama(self, nama):
        nama = nama.lower().replace(',', '.')
        nama = re.sub(r'\(.*?\)', '', nama)
        nama = re.sub(r'[\d]+(?:\.\d+)?', '', nama)
        nama = re.sub(r'[^a-z\s]', '', nama)
        return re.sub(r'\s+', ' ', nama).strip()

    def buat_alias_otomatis(self):
        cur = self.conn.cursor()

        cur.execute("SELECT DISTINCT nm_bahan FROM PembelianDetail")
        all_bahan = [row[0] for row in cur.fetchall()]
        
        norm_map = {bahan: self.normalisasi_nama(bahan) for bahan in all_bahan}

        grouped = defaultdict(list)
        used = set()

        for original, norm in norm_map.items():
            if original in used:
                continue
            group = [original]
            used.add(original)
            for other, other_norm in norm_map.items():
                if other in used:
                    continue
                ratio = difflib.SequenceMatcher(None, norm, other_norm).ratio()
                if ratio >= 0.95:
                    group.append(other)
                    used.add(other)
            nama_standar = min(group, key=lambda x: len(x))
            for alias in group:
                grouped[nama_standar].append(alias)

        for standar, aliases in grouped.items():
            for alias in aliases:
                cur.execute("REPLACE INTO BahanAlias (nama_lama, nama_standar) VALUES (?, ?)", (alias, standar))
        self.conn.commit()

    def get_stok_list_with_alias(self, limit=10, offset=0, search_query=None):
        cur = self.conn.cursor()

        # Ambil semua data stok
        if search_query:
            cur.execute("""
                SELECT S.nm_bahan, S.jumlah, S.satuan, S.harga_modal
                FROM Stok S
                WHERE LOWER(S.nm_bahan) LIKE ?
            """, (f"%{search_query.lower()}%",))
        else:
            cur.execute("SELECT nm_bahan, jumlah, satuan, harga_modal FROM Stok")
        
        rows = cur.fetchall()

        # Gabung dengan alias
        stok_map = {}
        for nama, qty, satuan, harga in rows:
            # Cek alias
            cur.execute("SELECT nama_standar FROM BahanAlias WHERE nama_lama = ?", (nama,))
            alias_row = cur.fetchone()
            nama_standar = alias_row[0] if alias_row else nama

            key = (nama_standar, satuan)
            if key in stok_map:
                stok_map[key]['jumlah'] += qty
            else:
                stok_map[key] = {
                    'nama': nama_standar,
                    'jumlah': qty,
                    'satuan': satuan,
                    'harga': harga
                }

        # Paging
        all_data = list(stok_map.values())
        # ✅ Urutkan berdasarkan nama bahan
        all_data.sort(key=lambda x: x['nama'])  # ascending
        paged = all_data[offset:offset + limit]

        return paged

    def count_stok_with_alias(self, search_query=None):
        data = self.get_stok_list_with_alias(limit=9999, offset=0, search_query=search_query)
        return len(data)
    
    def get_alias_list(self, nama_standar):
        cur = self.conn.cursor()
        rows = cur.execute("SELECT nama_lama FROM BahanAlias WHERE nama_standar = ?", (nama_standar,)).fetchall()
        return [row[0] for row in rows]

    def get_info_stok_by_satuan(self, nm_bahan):
        semua_data = self.get_info_stok(nm_bahan)
        
        # Ambil semua nama lama dari alias jika ada
        asal_alias = self.cursor.execute("""
            SELECT nama_lama FROM BahanAlias WHERE nama_standar = ?
        """, (nm_bahan,)).fetchall()
        asal_list = [row[0] for row in asal_alias] if asal_alias else []
        asal_list.append(nm_bahan)  # Tambahkan nama standar sendiri juga

        # Grouping berdasarkan satuan (tanpa case-sensitive)
        grouped = defaultdict(list)
        for row in semua_data:
            satuan_normal = row['satuan'].lower()
            row['asal_nama'] = self.get_asal_nama(row['no_nota'], row['tanggal'], row['jumlah'])
            grouped[satuan_normal.upper()].append(row)  # Tampilkan tetap huruf besar

        return grouped

    def get_asal_nama(self, no_nota, tanggal, jumlah):
        row = self.cursor.execute("""
            SELECT nm_bahan FROM PembelianDetail
            WHERE no_nota = ? AND jumlah = ? LIMIT 1
        """, (no_nota, jumlah)).fetchone()
        return row[0] if row else "-"


# Tambahkan di dalam kelas RestoDB (models.py)
    def hitung_rugi_laba(self, bulan):
        try:
            # Pendapatan dari penjualan
            self.cursor.execute("""
                SELECT SUM(d.jumlah * d.harga_satuan)
                FROM TransaksiDetail d
                JOIN Transaksi t ON d.no_nota = t.no_nota
                WHERE strftime('%Y-%m', t.tanggal) = ?
            """, (bulan,))
            pendapatan = self.cursor.fetchone()[0] or 0

            # HPP dari kebutuhan bahan
            self.cursor.execute("""
                SELECT SUM(k.jumlah * k.harga)
                FROM KebutuhanPaket k
                JOIN Transaksi t ON k.no_nota = t.no_nota
                WHERE strftime('%Y-%m', t.tanggal) = ?
            """, (bulan,))
            hpp = self.cursor.fetchone()[0] or 0

            laba_kotor = pendapatan - hpp

            # BIAYA OPERASIONAL dari tabel
            biaya_operasional = self.get_biaya_operasional_per_bulan(bulan)

            laba_bersih = laba_kotor - biaya_operasional

            # Simpan ke tabel laporan
            self.cursor.execute("""
                INSERT INTO NeracaRugiLaba (bulan, pendapatan, hpp, laba_kotor, biaya_operasional, laba_bersih)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (bulan, pendapatan, hpp, laba_kotor, biaya_operasional, laba_bersih))
            self.conn.commit()

            return {
                'bulan': bulan,
                'pendapatan': pendapatan,
                'hpp': hpp,
                'laba_kotor': laba_kotor,
                'biaya_operasional': biaya_operasional,
                'laba_bersih': laba_bersih
            }

        except Exception as e:
            print(f"Error menghitung rugi laba: {e}")
            return None


    def get_neraca_rugi_laba(self, bulan=None):
        if bulan:
            self.cursor.execute("""
                SELECT * FROM NeracaRugiLaba WHERE bulan = ? ORDER BY bulan DESC
            """, (bulan,))
        else:
            self.cursor.execute("""
                SELECT * FROM NeracaRugiLaba ORDER BY bulan DESC
            """)
        
        kolom = [desc[0] for desc in self.cursor.description]  # ambil nama kolom
        rows = self.cursor.fetchall()
        return [dict(zip(kolom, row)) for row in rows]  # convert ke dict

    def tambah_biaya_operasional(self, bulan, keterangan, jumlah):
        self.cursor.execute("""
            INSERT INTO BiayaOperasional (bulan, keterangan, jumlah)
            VALUES (?, ?, ?)
        """, (bulan, keterangan, jumlah))
        self.conn.commit()

    def get_biaya_operasional_per_bulan(self, bulan):
        self.cursor.execute("""
            SELECT SUM(jumlah)
            FROM BiayaOperasional
            WHERE strftime('%Y-%m', tanggal_pengeluaran) = ?
        """, (bulan,))
        return self.cursor.fetchone()[0] or 0

    def hitung_arus_kas(self, start_date, end_date):
        # --- Kas Masuk dari Transaksi Pembayaran TUNAI, TRANSFER, QRIS ---
        self.cursor.execute("""
            SELECT SUM(td.jumlah * td.harga_satuan)
            FROM TransaksiDetail td
            JOIN Transaksi t ON td.no_nota = t.no_nota
            WHERE DATE(t.tanggal) BETWEEN ? AND ?
            AND LOWER(t.jenis_pembayaran) IN ('tunai', 'transfer', 'qris')
        """, (start_date, end_date))
        kas_masuk = self.cursor.fetchone()[0] or 0

        # --- Kas Masuk dari Pelunasan Piutang ---
        self.cursor.execute("""
            SELECT SUM(jumlah)
            FROM PelunasanPiutang
            WHERE DATE(tanggal) BETWEEN ? AND ?
        """, (start_date, end_date))
        pelunasan_piutang = self.cursor.fetchone()[0] or 0

        # --- Kas Keluar: Biaya Operasional ---
        self.cursor.execute("""
            SELECT SUM(jumlah)
            FROM BiayaOperasional
            WHERE DATE(tanggal_pengeluaran) BETWEEN ? AND ?
        """, (start_date, end_date))
        biaya_operasional = self.cursor.fetchone()[0] or 0

        # --- Kas Keluar: Pembelian LUNAS ---
        self.cursor.execute("""
            SELECT SUM(pd.jumlah * pd.harga)
            FROM PembelianDetail pd
            JOIN Pembelian p ON pd.no_nota = p.no_nota
            WHERE DATE(p.tanggal) BETWEEN ? AND ?
            AND LOWER(p.status_pembayaran) = 'lunas'
        """, (start_date, end_date))
        pembelian_lunas = self.cursor.fetchone()[0] or 0

        # --- Kas Keluar: Pelunasan Utang ---
        self.cursor.execute("""
            SELECT SUM(jumlah)
            FROM PelunasanHutang
            WHERE DATE(tanggal) BETWEEN ? AND ?
        """, (start_date, end_date))
        pelunasan_utang = self.cursor.fetchone()[0] or 0

        # Total dan saldo kas
        kas_keluar = biaya_operasional + pembelian_lunas + pelunasan_utang
        kas_bersih = kas_masuk + pelunasan_piutang - kas_keluar

        return {
            'kas_masuk': kas_masuk,
            'pelunasan_piutang': pelunasan_piutang,
            'biaya_operasional': biaya_operasional,
            'pembelian_lunas': pembelian_lunas,
            'pelunasan_utang': pelunasan_utang,
            'kas_keluar': kas_keluar,
            'kas_bersih': kas_bersih
        }


    def ambil_detail_kas_masuk(self, start_date, end_date):
        self.cursor.execute("""
            SELECT t.tanggal, t.no_nota, t.jenis_pembayaran, SUM(td.jumlah * td.harga_satuan) as total
            FROM TransaksiDetail td
            JOIN Transaksi t ON td.no_nota = t.no_nota
            WHERE DATE(t.tanggal) BETWEEN ? AND ?
            AND LOWER(t.jenis_pembayaran) IN ('tunai', 'transfer', 'qris')
            GROUP BY t.no_nota
            ORDER BY t.tanggal ASC
        """, (start_date, end_date))
        return self.cursor.fetchall()

    def ambil_detail_biaya_operasional(self, start_date, end_date):
        self.cursor.execute("""
            SELECT tanggal_pengeluaran, keterangan, jumlah
            FROM BiayaOperasional
            WHERE DATE(tanggal_pengeluaran) BETWEEN ? AND ?
            ORDER BY tanggal_pengeluaran ASC
        """, (start_date, end_date))
        return self.cursor.fetchall()

    def ambil_detail_pembelian_lunas(self, start_date, end_date):
        self.cursor.execute("""
            SELECT p.tanggal, p.no_nota, SUM(pd.jumlah * pd.harga) as total
            FROM PembelianDetail pd
            JOIN Pembelian p ON pd.no_nota = p.no_nota
            WHERE DATE(p.tanggal) BETWEEN ? AND ?
            AND p.status_pembayaran = 'LUNAS'
            GROUP BY p.no_nota
            ORDER BY p.tanggal ASC
        """, (start_date, end_date))
        return self.cursor.fetchall()

    def ambil_detail_pelunasan_piutang(self, start_date, end_date):
        self.cursor.execute("""
            SELECT tanggal, no_nota, jumlah, keterangan, metode_pembayaran
            FROM PelunasanPiutang
            WHERE DATE(tanggal) BETWEEN ? AND ?
            ORDER BY tanggal ASC
        """, (start_date, end_date))
        return self.cursor.fetchall()

    def ambil_detail_pelunasan_utang(self, start_date, end_date):
        self.cursor.execute("""
            SELECT tanggal, no_nota, jumlah, keterangan
            FROM PelunasanHutang
            WHERE DATE(tanggal) BETWEEN ? AND ?
            ORDER BY tanggal ASC
        """, (start_date, end_date))
        return self.cursor.fetchall()

    def get_stok_saat_ini(self):
        self.cursor.execute("""
            SELECT nm_bahan, jumlah, satuan, harga_modal,
                (jumlah * harga_modal) as total_nilai
            FROM Stok
            ORDER BY nm_bahan COLLATE NOCASE ASC
        """)
        return self.cursor.fetchall()

    def get_histori_stok(self, start_date, end_date):
        # Barang masuk (pembelian)
        self.cursor.execute("""
            SELECT p.tanggal, d.nm_bahan, 'MASUK' as jenis, d.jumlah, d.satuan, 'Pembelian' as keterangan
            FROM PembelianDetail d
            JOIN Pembelian p ON d.no_nota = p.no_nota
            WHERE DATE(p.tanggal) BETWEEN ? AND ?
        """, (start_date, end_date))
        masuk = self.cursor.fetchall()

        # Barang keluar (pemakaian)
        self.cursor.execute("""
            SELECT t.tanggal, k.nm_bahan, 'KELUAR' as jenis, k.jumlah, k.satuan_pakai, 'Pemakaian Menu' as keterangan
            FROM KebutuhanPaket k
            JOIN Transaksi t ON k.no_nota = t.no_nota
            WHERE DATE(t.tanggal) BETWEEN ? AND ?
        """, (start_date, end_date))

        keluar = self.cursor.fetchall()

        hasil = masuk + keluar
        hasil.sort(key=lambda x: x[0])  # Urutkan berdasarkan tanggal
        return hasil

    def get_pergerakan_bahan(self, nama_bahan, start_date, end_date):
        # Data masuk (Pembelian)
        self.cursor.execute("""
            SELECT DATE(p.tanggal) as tgl, SUM(d.jumlah)
            FROM PembelianDetail d
            JOIN Pembelian p ON d.no_nota = p.no_nota
            WHERE d.nm_bahan = ? AND DATE(p.tanggal) BETWEEN ? AND ?
            GROUP BY DATE(p.tanggal)
        """, (nama_bahan, start_date, end_date))
        masuk = {row[0]: row[1] for row in self.cursor.fetchall()}

        # Data keluar (Pemakaian)
        self.cursor.execute("""
            SELECT DATE(t.tanggal) as tgl, SUM(k.jumlah)
            FROM KebutuhanPaket k
            JOIN Transaksi t ON k.no_nota = t.no_nota
            WHERE k.nm_bahan = ? AND DATE(t.tanggal) BETWEEN ? AND ?
            GROUP BY DATE(t.tanggal)
        """, (nama_bahan, start_date, end_date))
        keluar = {row[0]: row[1] for row in self.cursor.fetchall()}

        # Gabungkan tanggal
        all_dates = sorted(set(list(masuk.keys()) + list(keluar.keys())))

        result = []
        for tgl in all_dates:
            result.append({
                'tanggal': tgl,
                'masuk': masuk.get(tgl, 0),
                'keluar': keluar.get(tgl, 0)
            })
        return result

    def ambil_nama_bahan(self):
        self.cursor.execute("SELECT DISTINCT nm_bahan FROM PembelianDetail ORDER BY nm_bahan")
        return [row[0] for row in self.cursor.fetchall()]


    def hitung_pendapatan_hari_ini(self, tgl):
        self.cursor.execute("""
            SELECT SUM(jumlah * harga_satuan)
            FROM TransaksiDetail td
            JOIN Transaksi t ON td.no_nota = t.no_nota
            WHERE DATE(t.tanggal) = ?
        """, (tgl,))
        return self.cursor.fetchone()[0] or 0

    def hitung_pengeluaran_hari_ini(self, tgl):
        self.cursor.execute("""
            SELECT COALESCE((
                SELECT SUM(jumlah) FROM BiayaOperasional WHERE DATE(tanggal_pengeluaran) = ?
            ), 0) +
            COALESCE((
                SELECT SUM(jumlah * harga)
                FROM PembelianDetail pd
                JOIN Pembelian p ON pd.no_nota = p.no_nota
                WHERE DATE(p.tanggal) = ?
            ), 0)
        """, (tgl, tgl))
        return self.cursor.fetchone()[0] or 0

    def hitung_nilai_modal_stok(self):
        self.cursor.execute("""
            SELECT SUM(jumlah * IFNULL(harga_modal, 0)) FROM Stok
        """)
        return self.cursor.fetchone()[0] or 0

    def get_rugi_laba_per_bulan(self):
        self.cursor.execute("""
            SELECT bulan, laba_bersih FROM NeracaRugiLaba ORDER BY bulan ASC
        """)
        return self.cursor.fetchall()

    def get_arus_kas_per_bulan(self):
        self.cursor.execute("""
            SELECT strftime('%Y-%m', tanggal_pengeluaran), SUM(jumlah)
            FROM BiayaOperasional
            GROUP BY strftime('%Y-%m', tanggal_pengeluaran)
        """)
        return self.cursor.fetchall()

    def get_transaksi_terbaru(self, limit=5):
        self.cursor.execute("""
            SELECT tanggal, no_nota, jenis_pembayaran
            FROM Transaksi ORDER BY tanggal DESC LIMIT ?
        """, (limit,))
        return self.cursor.fetchall()

    def hitung_pendapatan_range(self, start, end):
        self.cursor.execute("""
            SELECT SUM(jumlah * harga_satuan)
            FROM TransaksiDetail td
            JOIN Transaksi t ON td.no_nota = t.no_nota
            WHERE DATE(t.tanggal) BETWEEN ? AND ?
        """, (start, end))
        return self.cursor.fetchone()[0] or 0

    def hitung_pengeluaran_range(self, start, end):
        self.cursor.execute("""
            SELECT COALESCE((
                SELECT SUM(jumlah)
                FROM BiayaOperasional
                WHERE DATE(tanggal_pengeluaran) BETWEEN ? AND ?
            ), 0) +
            COALESCE((
                SELECT SUM(jumlah * harga)
                FROM PembelianDetail pd
                JOIN Pembelian p ON pd.no_nota = p.no_nota
                WHERE DATE(p.tanggal) BETWEEN ? AND ?
            ), 0)
        """, (start, end, start, end))
        return self.cursor.fetchone()[0] or 0

    def get_jurnal_umum(self, start_date, end_date):
        data = []

        # Transaksi penjualan
        self.cursor.execute("""
            SELECT t.tanggal, t.no_nota, t.jenis_pembayaran, SUM(td.jumlah * td.harga_satuan) as total
            FROM Transaksi t
            JOIN TransaksiDetail td ON t.no_nota = td.no_nota
            WHERE DATE(t.tanggal) BETWEEN ? AND ?
            GROUP BY t.no_nota
        """, (start_date, end_date))

        for row in self.cursor.fetchall():
            tanggal, no_nota, metode, total = row
            akun_debet = 'Kas'
            kode_debet = '1000'

            if metode.upper() == 'BELUM':
                akun_debet = 'Piutang Usaha'
                kode_debet = '1100'

            data.append({
                'tanggal': tanggal,
                'akun': 'Pendapatan Usaha',
                'kode_akun': '4000',
                'debet': '',
                'kredit': total,
                'no_ref': no_nota,
                'keterangan': 'Penjualan'
            })
            data.append({
                'tanggal': tanggal,
                'akun': akun_debet,
                'kode_akun': kode_debet,
                'debet': total,
                'kredit': '',
                'no_ref': no_nota,
                'keterangan': 'Penjualan'
            })

        # Pembelian bahan baku
        self.cursor.execute("""
            SELECT p.tanggal, p.no_nota, SUM(d.jumlah * d.harga) as total
            FROM Pembelian p
            JOIN PembelianDetail d ON p.no_nota = d.no_nota
            WHERE DATE(p.tanggal) BETWEEN ? AND ?
            GROUP BY p.no_nota
        """, (start_date, end_date))
        for row in self.cursor.fetchall():
            tanggal, no_nota, total = row
            data.append({
                'tanggal': tanggal,
                'akun': 'Persediaan Bahan',
                'kode_akun': '5000',
                'debet': total,
                'kredit': '',
                'no_ref': no_nota,
                'keterangan': 'Pembelian bahan baku'
            })
            data.append({
                'tanggal': tanggal,
                'akun': 'Kas',
                'kode_akun': '1000',
                'debet': '',
                'kredit': total,
                'no_ref': no_nota,
                'keterangan': 'Pembelian bahan baku'
            })

        # Biaya operasional
        self.cursor.execute("""
            SELECT id, tanggal_pengeluaran, keterangan, jumlah
            FROM BiayaOperasional
            WHERE DATE(tanggal_pengeluaran) BETWEEN ? AND ?
        """, (start_date, end_date))
        for row in self.cursor.fetchall():
            id_biaya, tanggal, ket, jumlah = row
            no_ref = f"BO{id_biaya:03d}"
            data.append({
                'tanggal': tanggal,
                'akun': 'Biaya Operasional',
                'kode_akun': '6000',
                'debet': jumlah,
                'kredit': '',
                'no_ref': no_ref,
                'keterangan': ket
            })
            data.append({
                'tanggal': tanggal,
                'akun': 'Kas',
                'kode_akun': '1000',
                'debet': '',
                'kredit': jumlah,
                'no_ref': no_ref,
                'keterangan': ket
            })

        # Pelunasan Piutang
        self.cursor.execute("""
            SELECT id, tanggal, no_nota, jumlah, keterangan, metode_pembayaran
            FROM PelunasanPiutang
            WHERE DATE(tanggal) BETWEEN ? AND ?
        """, (start_date, end_date))

        for row in self.cursor.fetchall():
            id_pelunasan, tanggal, no_nota, jumlah, ket, metode = row
            no_ref = f"PL{str(id_pelunasan).zfill(4)}"
            keterangan_lengkap = f"{ket or 'Pelunasan Piutang'} - {metode.upper()}"
            
            data.append({
                'tanggal': tanggal,
                'akun': 'Kas',
                'kode_akun': '1000',
                'debet': jumlah,
                'kredit': '',
                'no_ref': no_ref,
                'keterangan': keterangan_lengkap
            })
            data.append({
                'tanggal': tanggal,
                'akun': 'Piutang Usaha',
                'kode_akun': '1100',
                'debet': '',
                'kredit': jumlah,
                'no_ref': no_ref,
                'keterangan': keterangan_lengkap
            })


        # Urutkan berdasarkan tanggal
        data.sort(key=lambda x: x['tanggal'])
        return data

    def get_buku_besar(self, kode_akun, start_date, end_date):
        jurnal = self.get_jurnal_umum(start_date, end_date)
        saldo = 0
        buku_besar = []

        for j in jurnal:
            if j['kode_akun'] != kode_akun:
                continue

            debet = j['debet'] or 0
            kredit = j['kredit'] or 0
            saldo += debet - kredit

            buku_besar.append({
                'tanggal': j['tanggal'],
                'keterangan': j['keterangan'],
                'debet': debet,
                'kredit': kredit,
                'saldo': saldo
            })

        return buku_besar

    def get_buku_besar_semua(self, start_date, end_date):
        jurnal = self.get_jurnal_umum(start_date, end_date)
        buku_besar_dict = {}

        # Kelompokkan berdasarkan kode akun
        for j in jurnal:
            kode = j['kode_akun']
            akun = j['akun']
            debet = j['debet'] or 0
            kredit = j['kredit'] or 0

            if kode not in buku_besar_dict:
                buku_besar_dict[kode] = {
                    'nama_akun': akun,
                    'entri': [],
                    'saldo': 0
                }

            # Hitung saldo berjalan
            saldo_sekarang = buku_besar_dict[kode]['saldo'] + debet - kredit
            buku_besar_dict[kode]['entri'].append({
                'tanggal': j['tanggal'],
                'keterangan': j['keterangan'],
                'debet': debet,
                'kredit': kredit,
                'saldo': saldo_sekarang
            })
            buku_besar_dict[kode]['saldo'] = saldo_sekarang

        return buku_besar_dict

   